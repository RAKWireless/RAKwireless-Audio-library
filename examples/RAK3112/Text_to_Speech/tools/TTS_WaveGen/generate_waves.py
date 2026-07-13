#!/usr/bin/env python3
"""
Generate WAV files for the WisBlock TTS library using Piper.

Reads an Excel sheet listing every WAV to produce, synthesises each line with
Piper, resamples to 22.05 kHz / 16-bit / mono PCM (the format the full TTS app
expects), and writes the file to output/<folder>/<filename>.

See README.md for the Excel schema and design decisions.
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
import wave
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
from openpyxl import Workbook, load_workbook
from piper import PiperVoice
from scipy.signal import resample_poly

DEFAULT_VOICE = "en_GB-alan-medium"
TARGET_SAMPLE_RATE = 22050  # 22.05 kHz - rate the full TTS app expects
TARGET_SAMPLE_WIDTH = 2  # bytes -> 16-bit
TARGET_CHANNELS = 1

SCRIPT_DIR = Path(__file__).resolve().parent
VOICES_DIR = SCRIPT_DIR / "voices"
OUTPUT_DIR = SCRIPT_DIR / "tts"
WORDLIST_XLSX = SCRIPT_DIR / "wordlist.xlsx"


def find_onnx(voice_name: str) -> Path | None:
    """Locate <voice_name>.onnx anywhere under VOICES_DIR (piper may nest it)."""
    if not VOICES_DIR.exists():
        return None
    matches = list(VOICES_DIR.rglob(f"{voice_name}.onnx"))
    return matches[0] if matches else None


def ensure_voice(voice_name: str) -> PiperVoice:
    """Download the Piper voice if missing, then load it."""
    VOICES_DIR.mkdir(parents=True, exist_ok=True)
    onnx = find_onnx(voice_name)
    if onnx is None:
        print(f"  downloading voice: {voice_name}")
        result = subprocess.run(
            [sys.executable, "-m", "piper.download_voices",
             voice_name, "--data-dir", str(VOICES_DIR)],
            capture_output=True, text=True,
        )
        if result.returncode != 0:
            tail = (result.stderr or result.stdout).strip().splitlines()[-3:]
            hint = ""
            if "404" in (result.stderr + result.stdout):
                hint = (
                    f"\nthis usually means the voice name doesn't exist in "
                    f"piper's repo. browse available voices at:\n"
                    f"  https://huggingface.co/rhasspy/piper-voices/tree/main\n"
                    f"voice names follow <lang>-<speaker>-<quality>, e.g. "
                    f"en_GB-cori-high, en_GB-alan-medium."
                )
            raise SystemExit(
                f"failed to download voice {voice_name!r}:\n  "
                + "\n  ".join(tail) + hint
            )
        onnx = find_onnx(voice_name)
        if onnx is None:
            raise SystemExit(
                f"download of {voice_name!r} reported success but no "
                f"{voice_name}.onnx was found under {VOICES_DIR}"
            )
    print(f"  voice ready: {voice_name}  ({onnx.relative_to(VOICES_DIR)})")
    return PiperVoice.load(str(onnx))


def collect_voices_needed(entries, default_voice: str) -> list[str]:
    """Unique voice names referenced by the wordlist, in first-seen order.
    Silence-sentinel rows don't need a voice and are skipped."""
    seen: dict[str, None] = {}
    for e in entries:
        if silence_samples(e["text"]) is not None:
            continue
        name = e["voice"] or default_voice
        seen.setdefault(name, None)
    return list(seen.keys())


_SILENCE_RE = re.compile(r"^<silence:(\d+)>$")


def silence_samples(text: str) -> np.ndarray | None:
    """If text is a silence sentinel like '<silence:100>', return that many
    ms of zero int16 samples at TARGET_SAMPLE_RATE. Otherwise None."""
    m = _SILENCE_RE.match(text.strip())
    if not m:
        return None
    ms = int(m.group(1))
    n = (TARGET_SAMPLE_RATE * ms) // 1000
    return np.zeros(n, dtype=np.int16)


def synth_to_array(voice: PiperVoice, text: str) -> tuple[np.ndarray, int]:
    """Run Piper, return mono int16 samples and the model's sample rate."""
    chunks = list(voice.synthesize(text))
    if not chunks:
        raise RuntimeError(f"Piper returned no audio for: {text!r}")
    sr = chunks[0].sample_rate
    pcm = b"".join(c.audio_int16_bytes for c in chunks)
    samples = np.frombuffer(pcm, dtype=np.int16)
    return samples, sr


def to_target_format(samples: np.ndarray, src_sr: int) -> np.ndarray:
    """Resample to 16 kHz mono int16."""
    if src_sr == TARGET_SAMPLE_RATE:
        return samples
    # resample_poly works in float; cast back to int16 with clipping.
    from math import gcd
    g = gcd(src_sr, TARGET_SAMPLE_RATE)
    up = TARGET_SAMPLE_RATE // g
    down = src_sr // g
    resampled = resample_poly(samples.astype(np.float32), up, down)
    return np.clip(resampled, -32768, 32767).astype(np.int16)


def write_wav(path: Path, samples: np.ndarray) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with wave.open(str(path), "wb") as w:
        w.setnchannels(TARGET_CHANNELS)
        w.setsampwidth(TARGET_SAMPLE_WIDTH)
        w.setframerate(TARGET_SAMPLE_RATE)
        w.writeframes(samples.tobytes())


def clean_output(out_root: Path) -> int:
    """Delete the regenerable, gitignored artifacts:
      - the generated output tree (all WAVs + manifest.json)
      - the downloaded Piper voice models (voices/)

    Only files are removed; the directory structure is left in place. Both
    sets of files are reproducible (the library from wordlist.xlsx, the voices
    by re-download on the next run) and are gitignored, so they should never be
    committed - run this (`--clean`) to clear them before pushing to git.
    Returns the total number of files removed.
    """
    total = 0
    for target, label in ((out_root, "output"), (VOICES_DIR, "voices")):
        if not target.exists():
            print(f"nothing to clean ({label}): {target} does not exist")
            continue
        files = [p for p in target.rglob("*") if p.is_file()]
        for p in files:
            p.unlink()
        total += len(files)
        print(f"removed {len(files)} {label} file(s) "
              f"(directories kept) under: {target}")
    return total


def dedupe_xlsx(xlsx_path: Path) -> int:
    """Remove duplicate rows from the wordlist, in place.

    A duplicate is any row whose (folder, filename) pair -- i.e. its output
    WAV path -- matches an earlier row. The first occurrence is kept; later
    ones are deleted and the workbook is saved. Returns the number removed.
    """
    wb = load_workbook(xlsx_path)  # writable (not read_only)
    ws = wb.active
    header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    required = {"text", "filename", "folder"}
    if not required.issubset(header):
        raise SystemExit(
            f"{xlsx_path.name} must have columns: text, filename, folder "
            f"(optional: voice). Found: {header}"
        )
    fcol = header.index("folder")
    ncol = header.index("filename")

    seen: dict[tuple[str, str], int] = {}
    dup_rows: list[tuple[int, str, str]] = []
    for r in range(2, ws.max_row + 1):
        folder = ws.cell(row=r, column=fcol + 1).value
        filename = ws.cell(row=r, column=ncol + 1).value
        if folder is None or filename is None:
            continue
        key = (str(folder).strip().strip("/"), str(filename).strip())
        if key in seen:
            dup_rows.append((r, key[0], key[1]))
        else:
            seen[key] = r

    if not dup_rows:
        return 0

    print(f"  removing {len(dup_rows)} duplicate row(s) from {xlsx_path.name}:")
    for r, folder, filename in dup_rows:
        print(f"    row {r}: {folder}/{filename} "
              f"(first seen at row {seen[(folder, filename)]})")
    # Delete from the bottom up so earlier row indices stay valid.
    for r, _, _ in sorted(dup_rows, reverse=True):
        ws.delete_rows(r, 1)
    wb.save(xlsx_path)
    return len(dup_rows)


def iter_rows(xlsx_path: Path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip().lower() if c else "" for c in next(rows)]
    required = {"text", "filename", "folder"}
    if not required.issubset(header):
        raise SystemExit(
            f"{xlsx_path.name} must have columns: text, filename, folder "
            f"(optional: voice). Found: {header}"
        )
    idx = {name: header.index(name) for name in header if name}
    for i, row in enumerate(rows, start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        def get(name):
            j = idx.get(name)
            return row[j] if j is not None and j < len(row) else None
        text = get("text")
        filename = get("filename")
        folder = get("folder")
        voice = get("voice")
        if not text or not filename or folder is None:
            print(f"  row {i}: skipped (missing text/filename/folder)")
            continue
        yield {
            "row": i,
            "text": str(text).strip(),
            "filename": str(filename).strip(),
            "folder": str(folder).strip().strip("/"),
            "voice": str(voice).strip() if voice else None,
        }


def init_template(path: Path) -> None:
    """Write a starter wordlist.xlsx with a handful of example rows."""
    if path.exists():
        raise SystemExit(f"refusing to overwrite existing {path.name}")
    wb = Workbook()
    ws = wb.active
    ws.title = "wordlist"
    ws.append(["text", "filename", "folder", "voice"])
    examples = [
        ("Battery is low.",         "battery_low.wav",       "phrases",       ""),
        ("Message received.",       "message_received.wav",  "phrases",       ""),
        ("Temperature is",          "temperature_is.wav",    "phrases",       ""),
        ("Hello.",                  "hello.wav",             "phrases",       ""),
        ("zero",                    "0.wav",                 "words/numbers", ""),
        ("one",                     "1.wav",                 "words/numbers", ""),
        ("two",                     "2.wav",                 "words/numbers", ""),
        ("point",                   "point.wav",             "words/numbers", ""),
        ("degrees",                 "degrees.wav",           "words/units",   ""),
        ("celsius",                 "celsius.wav",           "words/units",   ""),
        ("battery",                 "battery.wav",           "words/nouns",   ""),
        # voice override demo: alan only ships in low/medium, so use the
        # 'cori' speaker for a high-quality UK English comparison.
        ("Hello, this is the high quality voice.",
                                    "hello_high.wav",        "phrases",       "en_GB-cori-high"),
    ]
    for row in examples:
        ws.append(row)
    for col, width in zip("ABCD", (40, 28, 18, 22)):
        ws.column_dimensions[col].width = width
    wb.save(path)
    print(f"wrote template: {path}")


_PUNCT_RE = re.compile(r"[^\w\s]+", re.UNICODE)
_WS_RE = re.compile(r"\s+")


def normalize_text(text: str) -> str:
    """Lowercase, strip punctuation, collapse whitespace.

    The on-device matcher must apply the same rule to incoming text before
    looking it up against this manifest.
    """
    return _WS_RE.sub(" ", _PUNCT_RE.sub(" ", text.lower())).strip()


def write_manifest(entries, out_root: Path) -> Path:
    """Emit output/manifest.json. Schema documented in README.md."""
    phrases, words = [], []
    for e in entries:
        folder = e["folder"].strip("/")
        sd_path = f"/tts/{folder}/{e['filename']}"
        record = {"text": normalize_text(e["text"]), "file": sd_path}
        if folder == "phrases":
            phrases.append(record)
        elif folder.startswith("words/"):
            record["category"] = folder[len("words/"):]
            words.append(record)
        elif folder == "words":
            words.append(record)
        # other folders (e.g. silence/) are intentionally not in the lookup
        # tiers; the device addresses them directly by path.

    manifest = {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "format": {
            "sample_rate": TARGET_SAMPLE_RATE,
            "bits": TARGET_SAMPLE_WIDTH * 8,
            "channels": TARGET_CHANNELS,
        },
        "phrases": phrases,
        "words": words,
    }
    out_root.mkdir(parents=True, exist_ok=True)
    path = out_root / "manifest.json"
    path.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    return path


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", type=Path, default=WORDLIST_XLSX,
                    help="path to wordlist xlsx (default: wordlist.xlsx)")
    ap.add_argument("--out", type=Path, default=OUTPUT_DIR,
                    help="output root directory (default: ./tts)")
    ap.add_argument("--default-voice", default=DEFAULT_VOICE,
                    help=f"voice when row's voice column is blank "
                         f"(default: {DEFAULT_VOICE})")
    ap.add_argument("--force", action="store_true",
                    help="regenerate WAVs even if they already exist")
    ap.add_argument("--init", action="store_true",
                    help="write a starter wordlist.xlsx and exit")
    ap.add_argument("--manifest-only", action="store_true",
                    help="rebuild manifest.json without (re)synthesising")
    ap.add_argument("--clean", action="store_true",
                    help="delete the output dir (generated WAVs + manifest) "
                         "and the downloaded voices, then exit. Both are "
                         "regenerable and gitignored - run before pushing to git.")
    args = ap.parse_args()

    if args.clean:
        clean_output(args.out)
        return 0

    if args.init:
        init_template(args.xlsx)
        return 0

    if not args.xlsx.exists():
        print(f"no wordlist found at {args.xlsx}")
        print("run with --init to create a starter template.")
        return 1

    # Strip duplicate (folder, filename) rows from the sheet before anything
    # else reads it. This rewrites wordlist.xlsx in place.
    removed = dedupe_xlsx(args.xlsx)
    if removed:
        print(f"de-duplicated wordlist: {removed} row(s) removed.\n")

    entries = list(iter_rows(args.xlsx))
    if not entries:
        print("no usable rows in wordlist.")
        return 1

    if args.manifest_only:
        path = write_manifest(entries, args.out)
        print(f"wrote manifest: {path}")
        return 0

    # Pre-flight: download + load every voice referenced by the sheet before
    # touching synthesis. Failing here is cheap; failing mid-run is annoying.
    needed = collect_voices_needed(entries, args.default_voice)
    print(f"voices required: {', '.join(needed)}")
    voice_cache: dict[str, PiperVoice] = {name: ensure_voice(name) for name in needed}
    print()

    made = skipped = 0
    for entry in entries:
        out_path = args.out / entry["folder"] / entry["filename"]
        if out_path.exists() and not args.force:
            skipped += 1
            continue

        sil = silence_samples(entry["text"])
        if sil is not None:
            print(f"  [silence] {entry['folder']}/{entry['filename']}  <- {entry['text']}")
            write_wav(out_path, sil)
            made += 1
            continue

        voice_name = entry["voice"] or args.default_voice
        voice = voice_cache[voice_name]
        print(f"  [{voice_name}] {entry['folder']}/{entry['filename']}  <- {entry['text']!r}")
        samples, src_sr = synth_to_array(voice, entry["text"])
        samples = to_target_format(samples, src_sr)
        write_wav(out_path, samples)
        made += 1

    manifest_path = write_manifest(entries, args.out)
    print(f"\ndone: {made} written, {skipped} skipped (already existed).")
    print(f"output root: {args.out}")
    print(f"manifest:    {manifest_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
